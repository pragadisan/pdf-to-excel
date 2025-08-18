import os, re, math, shutil
from decimal import Decimal, InvalidOperation
from dataclasses import dataclass
from typing import List, Optional, Tuple

import fitz  # PyMuPDF
from PIL import Image
import numpy as np
import easyocr

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows

import tkinter as tk
from tkinter import filedialog, messagebox

# =========================
# Parsing configuration
# =========================
# Strict money: require .00 style decimals (prevents "009" being captured)
MONEY_RX = re.compile(r'(?<!\d)(\d{1,3}(?:,\d{3})*|\d+)\.\d{2}(?!\d)')

# Dates: 01-06-2023, 1/6/23, 01.06.2023, 01 Jun 2023, 01-JUN-23, etc.
DATE_RXES = [
    re.compile(r'\b(\d{1,2}[-/\.]\d{1,2}[-/\.]\d{2,4})\b'),
    re.compile(r'\b(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4})\b'),
    re.compile(r'\b(\d{1,2}[-/\.][A-Za-z]{3}[-/\.]\d{2,4})\b'),
]

# Debit/Credit cues
CREDIT_HINTS = ("CR", "CREDIT", "DEPOSIT", "UPI/CR", "REFUND", "REVERSAL")
DEBIT_HINTS  = ("DR", "DEBIT", "WITHDRAWAL", "W/D", "ATM", "POS", "IRCTC", "CHARGES")

# Optional cheque extraction (not part of the final 5 columns; kept for future tuning)
CHEQUE_RXES = [
    re.compile(r'\bChq[:\s#]*([0-9]{4,})\b', re.IGNORECASE),
    re.compile(r'\bCheque[:\s#]*([0-9]{4,})\b', re.IGNORECASE),
    re.compile(r'\bCHQ[:\s#]*([0-9]{4,})\b', re.IGNORECASE),
]

DEBUG_DIR_NAME = "_debug_bank_ocr"  # created next to your source PDF
OCR_DPI = 300
EASYOCR_LANGS = ['en']  # add more like ['en','ta'] if you want

# =========================
# Utilities
# =========================
def normspaces(s: str) -> str:
    return re.sub(r'\s+', ' ', s).strip()

def clean_money(s: str) -> Decimal:
    return Decimal(s.replace(',', '').strip())

def find_date(s: str) -> str:
    for rx in DATE_RXES:
        m = rx.search(s)
        if m:
            return m.group(1)
    return ""

def has_any(s: str, words: Tuple[str, ...]) -> bool:
    up = s.upper()
    return any(w in up for w in words)

@dataclass
class ParsedRow:
    Date: str
    Particulars: str
    Debit: Decimal
    Credit: Decimal
    Balance: Decimal
    raw: str

# =========================
# LINE → ROW parser
# =========================
def parse_line(line: str, prev_balance: Optional[Decimal]) -> Optional[ParsedRow]:
    """
    Heuristic:
    - last money on the line = Balance (most bank layouts)
    - second-last money = movement Amount
    - decide Debit/Credit using keywords; else infer from running balance
    """
    raw = line
    line = normspaces(line)
    monies = [m.group(0) for m in MONEY_RX.finditer(line)]
    if len(monies) < 2:
        return None  # need at least [amount, balance]

    # pick balance and amount
    try:
        balance = clean_money(monies[-1])
        amount  = clean_money(monies[-2])
    except InvalidOperation:
        return None

    # detect/guess direction
    debit  = Decimal("0.00")
    credit = Decimal("0.00")

    if has_any(line, CREDIT_HINTS):
        credit = amount
    elif has_any(line, DEBIT_HINTS):
        debit = amount
    else:
        # infer using previous balance (if available)
        if prev_balance is not None:
            diff = balance - prev_balance
            # small tolerance for OCR/rounding
            if abs(diff - amount) <= Decimal("1.00"):
                credit = amount
            elif abs(diff + amount) <= Decimal("1.00"):
                debit = amount
            else:
                # fall back: if a minus sign right before amount, treat as debit
                # (common in some PDFs)
                before = line.rsplit(monies[-2], 1)[0][-3:]
                if "-" in before:
                    debit = amount
                else:
                    # default guess: credit
                    credit = amount
        else:
            # first row, no prev balance → guess from keywords around amount
            before = line.rsplit(monies[-2], 1)[0][-12:].upper()
            after  = line.split(monies[-2], 1)[1][:12].upper()
            context = before + " " + after
            if has_any(context, DEBIT_HINTS):
                debit = amount
            else:
                credit = amount

    # Particulars: strip trailing numbers and CR/DR tags
    tail = line
    tail = tail.rsplit(monies[-1], 1)[0]            # remove balance
    tail = tail.rsplit(monies[-2], 1)[0]            # remove amount
    tail = re.sub(r'(CR|DR)\s*$', '', tail, flags=re.IGNORECASE)
    particulars = normspaces(tail)

    # Date: pick from line or leave blank (some layouts keep date on a separate line)
    date = find_date(line)

    return ParsedRow(
        Date=date,
        Particulars=particulars,
        Debit=debit,
        Credit=credit,
        Balance=balance,
        raw=raw
    )

# =========================
# TEXT MODE (no OCR)
# =========================
def extract_text_lines_from_pdf(pdf_path: str) -> List[str]:
    lines: List[str] = []
    with fitz.open(pdf_path) as doc:
        for pno, page in enumerate(doc, 1):
            t = page.get_text("text") or ""
            for ln in t.splitlines():
                ln = normspaces(ln)
                if ln:
                    lines.append(ln)
    return lines

# =========================
# OCR MODE (EasyOCR CPU)
# =========================
def render_pdf_to_images(pdf_path: str, out_dir: str) -> List[str]:
    img_paths = []
    with fitz.open(pdf_path) as doc:
        for idx, page in enumerate(doc, 1):
            pix = page.get_pixmap(dpi=OCR_DPI)
            img_path = os.path.join(out_dir, f"page_{idx:03d}.png")
            pix.save(img_path)
            img_paths.append(img_path)
    return img_paths

def easyocr_lines(img_path: str, reader: easyocr.Reader) -> List[Tuple[float, float, str]]:
    """
    Returns list of (y_center, x_left, text) for each detected word,
    then we will group them into lines by y proximity and sort left→right.
    """
    img = np.array(Image.open(img_path).convert("RGB"))
    results = reader.readtext(img, detail=1, paragraph=False)  # ([box], text, conf)
    words = []
    for box, text, conf in results:
        # box is 4 points [[x1,y1],[x2,y2],[x3,y3],[x4,y4]]
        xs = [p[0] for p in box]
        ys = [p[1] for p in box]
        x_left = float(min(xs))
        y_center = float(sum(ys)/len(ys))
        words.append((y_center, x_left, text))
    # cluster words into lines by y
    words.sort(key=lambda t: (t[0], t[1]))
    lines: List[str] = []
    if not words:
        return lines

    current_y = words[0][0]
    current_line: List[Tuple[float, float, str]] = []
    y_threshold = 10.0  # pixels; adjust if needed

    for y, x, tx in words:
        if abs(y - current_y) <= y_threshold:
            current_line.append((y, x, tx))
        else:
            # flush previous line
            current_line.sort(key=lambda t: t[1])
            lines.append(normspaces(" ".join(w[2] for w in current_line)))
            # start new line
            current_line = [(y, x, tx)]
            current_y = y

    if current_line:
        current_line.sort(key=lambda t: t[1])
        lines.append(normspaces(" ".join(w[2] for w in current_line)))
    return lines

def ocr_pdf_to_lines(pdf_path: str, debug_dir: str) -> List[str]:
    os.makedirs(debug_dir, exist_ok=True)
    img_paths = render_pdf_to_images(pdf_path, debug_dir)
    reader = easyocr.Reader(EASYOCR_LANGS, gpu=False)
    all_lines: List[str] = []
    for p in img_paths:
        page_lines = easyocr_lines(p, reader)
        all_lines.extend(page_lines)
    # save for debugging
    with open(os.path.join(debug_dir, "ocr_lines.txt"), "w", encoding="utf-8") as f:
        f.write("\n".join(all_lines))
    return all_lines

# =========================
# MASTER parse workflow
# =========================
def parse_lines_to_rows(lines: List[str]) -> pd.DataFrame:
    rows: List[ParsedRow] = []
    prev_balance: Optional[Decimal] = None
    pending_desc: List[str] = []
    pending_date: str = ""

    for ln in lines:
        # collect date if the line is only a date or starts with date
        dt = find_date(ln)
        if dt and not MONEY_RX.search(ln):
            pending_date = dt
            # also add any preceding text to description buffer
            tail = normspaces(ln.replace(dt, "")).strip(" -–—")
            if tail:
                pending_desc.append(tail)
            continue

        # try parse as a transaction line (needs amounts)
        parsed = parse_line(ln, prev_balance)
        if parsed:
            # if parsed date missing, use pending_date
            if not parsed.Date and pending_date:
                parsed.Date = pending_date
            # prepend buffered description if any
            if pending_desc:
                parsed.Particulars = normspaces(" ".join(pending_desc) + " " + parsed.Particulars)
                pending_desc = []
            rows.append(parsed)
            prev_balance = parsed.Balance
        else:
            # accumulate description fragments between date and amount line
            if any(ch.isalpha() for ch in ln):  # avoid pure headers or numbers
                pending_desc.append(ln)

    if not rows:
        return pd.DataFrame(columns=["Date","Particulars","Debit","Credit","Balance"])

    df = pd.DataFrame([{
        "Date": r.Date,
        "Particulars": r.Particulars,
        "Debit": float(r.Debit),
        "Credit": float(r.Credit),
        "Balance": float(r.Balance),
        "raw": r.raw
    } for r in rows])

    # Clean: fill missing dates by forward-fill (some PDFs show date once then multiple lines)
    df["Date"] = df["Date"].replace("", pd.NA).ffill().fillna("")
    return df

def parse_pdf_to_dataframe(pdf_path: str, debug_root: str) -> pd.DataFrame:
    # 1) Try fast text mode
    text_lines = extract_text_lines_from_pdf(pdf_path)
    debug_dir = os.path.join(os.path.dirname(pdf_path), DEBUG_DIR_NAME)
    os.makedirs(debug_dir, exist_ok=True)
    with open(os.path.join(debug_dir, "text_lines.txt"), "w", encoding="utf-8") as f:
        f.write("\n".join(text_lines))

    df = parse_lines_to_rows(text_lines)
    if not df.empty:
        return df

    # 2) Fallback to OCR mode
    ocr_lines = ocr_pdf_to_lines(pdf_path, debug_dir)
    df = parse_lines_to_rows(ocr_lines)
    return df

# =========================
# Excel export with Table
# =========================
def save_dataframe_as_excel_table(df: pd.DataFrame, out_path: str, sheet_name="Statement"):
    # Only keep the requested 5 columns
    wanted_cols = ["Date","Particulars","Debit","Credit","Balance"]
    df = df[wanted_cols].copy()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write dataframe
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Convert range to an Excel Table (auto filter / banded rows)
    last_row = ws.max_row
    last_col = ws.max_column
    ref = f"A1:{chr(64+last_col)}{last_row}"
    tbl = Table(displayName="TxnTable", ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)

    # Column widths & number formats
    widths = [12, 60, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w
    money_fmt = '#,##0.00'
    for col in ('C','D','E'):
        for r in range(2, last_row+1):
            ws[f"{col}{r}"].number_format = money_fmt

    wb.save(out_path)

# =========================
# GUI
# =========================
class App:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("PDF → (Text/OCR) → Excel Table")
        self.pdf = None

        tk.Button(self.root, text="1) Pick PDF", command=self.pick).pack(fill="x", pady=4)
        tk.Button(self.root, text="2) Convert to Excel", command=self.convert).pack(fill="x", pady=4)

        self.txt = tk.Text(self.root, height=22)
        self.txt.pack(fill="both", expand=True)

    def log(self, s: str):
        self.txt.insert("end", s + "\n")
        self.txt.see("end")
        self.root.update()

    def pick(self):
        p = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if p:
            self.pdf = p
            self.txt.delete("1.0","end")
            self.log(f"PDF: {p}")
            self.log(f"Debug folder will be created next to PDF: {DEBUG_DIR_NAME}")

    def convert(self):
        if not self.pdf:
            messagebox.showwarning("Pick PDF", "Please pick a PDF first.")
            return

        self.log("Parsing (text mode) and falling back to OCR if needed…")
        try:
            df = parse_pdf_to_dataframe(self.pdf, DEBUG_DIR_NAME)
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return

        if df.empty:
            self.log("No rows parsed. Check _debug files for clues.")
            return

        # Save debug CSV preview
        debug_dir = os.path.join(os.path.dirname(self.pdf), DEBUG_DIR_NAME)
        os.makedirs(debug_dir, exist_ok=True)
        preview_csv = os.path.join(debug_dir, "parsed_preview.csv")
        df.to_csv(preview_csv, index=False, encoding="utf-8")
        self.log(f"Parsed {len(df)} rows. Preview CSV: {preview_csv}")

        out = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                           filetypes=[("Excel files","*.xlsx")],
                                           initialfile="statement.xlsx")
        if not out:
            return
        save_dataframe_as_excel_table(df, out)
        self.log(f"Excel saved → {out}")

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    App().run()
